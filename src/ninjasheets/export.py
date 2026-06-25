"""Export pipeline results to a reviewable .xlsx workbook.

Tabs follow the spec: Start Here, Runs, Videos, Transcript Raw, Athletes, Gyms,
Corrections, Processing Log. The Runs tab is the product: important columns
leftmost, frozen header, autofilter, clickable hyperlinks. No fake data --
empty enrichment tabs ship with headers only.
"""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .pipeline import RUN_COLUMNS, PipelineResult

# Runs columns surfaced first for the parent/coach workflow.
RUNS_DISPLAY_ORDER = [
    "athlete_name_clean",
    "gym_raw",
    "city",
    "state",
    "division",
    "gender",
    "tier",
    "stage",
    "wave",
    "run_order_overall",
    "timestamp_hhmmss",
    "youtube_run_url",
    "source_confidence",
    "review_status",
    "notes",
]
# Remaining (technical) columns appended to the right, original order preserved.
RUNS_COLUMN_ORDER = RUNS_DISPLAY_ORDER + [
    c for c in RUN_COLUMNS if c not in RUNS_DISPLAY_ORDER
]

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)
LINK_FONT = Font(color="2563EB", underline="single")

CONF_FILL = {
    "high": PatternFill("solid", fgColor="DCFCE7"),
    "medium": PatternFill("solid", fgColor="FEF9C3"),
    "low": PatternFill("solid", fgColor="FEE2E2"),
}

TRANSCRIPT_COLUMNS = [
    "video_id", "start_seconds", "duration_seconds", "end_seconds",
    "text_raw", "text_normalized",
]
ATHLETE_COLUMNS = [
    "athlete_id", "athlete_name", "athlete_name_normalized", "gym_raw",
    "gym_normalized", "city", "state", "country", "source_url", "source_type",
    "last_updated",
]
GYM_COLUMNS = [
    "gym_id", "gym_raw", "gym_normalized", "city", "state", "country",
    "source_url", "notes",
]
CORRECTION_COLUMNS = [
    "submitted_at", "submitter_name", "submitter_email", "run_id", "field",
    "current_value", "suggested_value", "evidence", "status", "reviewer_notes",
]
LOG_COLUMNS = [
    "run_datetime", "video_id", "step", "status", "rows_created", "warnings",
    "errors", "notes",
]


def _style_header(ws, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}1"


def _autosize(ws, max_width: int = 60) -> None:
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        longest = max((len(str(c.value)) for c in col_cells if c.value is not None),
                      default=10)
        ws.column_dimensions[letter].width = min(max(12, longest + 2), max_width)


def _write_table(ws, columns: list[str], rows: list[dict]) -> None:
    ws.append(columns)
    for r in rows:
        ws.append([r.get(c, "") for c in columns])
    _style_header(ws, len(columns))
    _autosize(ws)


def _build_start_here(ws, result: PipelineResult) -> None:
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    v = result.video
    rows = [
        ("ninjasheets", ""),
        ("A searchable, filterable index of World Ninja League runs.", ""),
        ("", ""),
        ("Coverage", f"{v.event_name} — {v.tier} | {v.division} {v.gender} | {v.stage}"),
        ("Video", v.video_url),
        ("Runs indexed", str(len(result.runs))),
        ("Last updated", now),
        ("", ""),
        ("How to use", ""),
        ("1.", "Open the Runs tab."),
        ("2.", "Search or filter by athlete, gym, division, tier, or stage."),
        ("3.", "Click the YouTube Run Link to jump straight to that run."),
        ("4.", "Submit fixes on the Corrections tab (or the linked form)."),
        ("", ""),
        ("Confidence", "high / medium / low reflect how sure the auto-extraction is."),
        ("Review status", "All auto rows ship as 'unreviewed' until a human checks them."),
        ("", ""),
        ("Disclaimer",
         "Unofficial community-created viewing index. Not affiliated with WNL. "
         "Links point to public YouTube videos. Submit corrections or removal "
         "requests via the Corrections tab."),
    ]
    ws.append(["ninjasheets", ""])
    ws["A1"].font = Font(size=18, bold=True, color="1F2937")
    for label, value in rows[1:]:
        ws.append([label, value])
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 90
    for row in ws.iter_rows():
        row[1].alignment = Alignment(wrap_text=True, vertical="top")


def _build_runs(ws, result: PipelineResult) -> None:
    cols = RUNS_COLUMN_ORDER
    ws.append(cols)
    url_idx = cols.index("youtube_run_url") + 1
    conf_idx = cols.index("source_confidence") + 1
    for r in result.runs:
        ws.append([r.get(c, "") for c in cols])
        row = ws.max_row
        link = ws.cell(row=row, column=url_idx)
        if link.value:
            link.hyperlink = link.value
            link.value = "▶ Watch run"
            link.font = LINK_FONT
        conf = str(ws.cell(row=row, column=conf_idx).value)
        fill = CONF_FILL.get(conf)
        if fill:
            ws.cell(row=row, column=conf_idx).fill = fill
    _style_header(ws, len(cols))
    _autosize(ws)
    ws.column_dimensions[get_column_letter(cols.index("detected_text") + 1)].width = 60
    ws.column_dimensions[get_column_letter(cols.index("notes") + 1)].width = 40


def _build_videos(ws, result: PipelineResult) -> None:
    v = result.video
    meta = result.metadata
    cols = [
        "video_id", "video_url", "video_title", "event_name", "season",
        "competition_name", "tier", "division", "gender", "stage", "course",
        "date", "duration_seconds", "transcript_available", "processing_status",
        "source_notes",
    ]
    row = {
        "video_id": v.video_id, "video_url": v.video_url,
        "video_title": v.video_title or meta.get("title", ""),
        "event_name": v.event_name, "season": v.season,
        "competition_name": v.competition_name, "tier": v.tier,
        "division": v.division, "gender": v.gender, "stage": v.stage,
        "course": v.course, "date": v.date,
        "duration_seconds": meta.get("duration", ""),
        "transcript_available": bool(result.transcript_lines),
        "processing_status": "processed" if result.runs else "needs_review",
        "source_notes": "Phase 1 transcript extraction (en-orig auto-captions).",
    }
    _write_table(ws, cols, [row])


def _build_transcript_raw(ws, result: PipelineResult) -> None:
    rows = [
        {
            "video_id": ln.video_id,
            "start_seconds": round(ln.start_seconds, 2),
            "duration_seconds": round(ln.duration_seconds, 2),
            "end_seconds": round(ln.end_seconds, 2),
            "text_raw": ln.text_raw,
            "text_normalized": ln.text_normalized,
        }
        for ln in result.transcript_lines
    ]
    _write_table(ws, TRANSCRIPT_COLUMNS, rows)


def build_workbook(result: PipelineResult, output_path: Path) -> Path:
    wb = Workbook()
    _build_start_here(wb.active, result)
    wb.active.title = "Start Here"

    _build_runs(wb.create_sheet("Runs"), result)
    _build_videos(wb.create_sheet("Videos"), result)
    _build_transcript_raw(wb.create_sheet("Transcript Raw"), result)
    _write_table(wb.create_sheet("Athletes"), ATHLETE_COLUMNS, [])
    _write_table(wb.create_sheet("Gyms"), GYM_COLUMNS, [])
    _write_table(wb.create_sheet("Corrections"), CORRECTION_COLUMNS, [])
    _write_table(wb.create_sheet("Processing Log"), LOG_COLUMNS, result.log)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
